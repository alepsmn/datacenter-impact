import openpyxl
wb = openpyxl.load_workbook("data/raw/epri/EPRI_2024_Projections.xlsx", read_only=True)
print(wb.sheetnames)

ws = wb.active
for i, row in enumerate(ws.iter_rows(values_only=True)):
    print(row)
    if i >= 5:
        break